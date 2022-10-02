import openpyxl
import random
book = openpyxl.load_workbook("Task 1.xlsx", read_only=True)
sheetname = book.worksheets[0]
sheet = book.worksheets[1]

variants = list()
for row in range(3, sheet.max_row + 1):
    variants.append(sheet[row][0].value)

t1y1 = list()
for row in range(3, sheet.max_row + 1):
    t1y1.append(sheet[row][1].value)

t1m1 = list()
for row in range(3, sheet.max_row + 1):
    t1m1.append(sheet[row][2].value)

t1y2 = list()
for row in range(3, sheet.max_row + 1):
    t1y2.append(sheet[row][3].value)

t1m2 = list()
for row in range(3, sheet.max_row + 1):
    t1m2.append(sheet[row][4].value)

t1v1 = list()
for row in range(3, sheet.max_row + 1):
    t1v1.append(sheet[row][5].value)

t1Tmoor = list()
for row in range(3, sheet.max_row + 1):
    t1Tmoor.append(sheet[row][6].value)

# Delta T
dt = list()
for i in range(len(variants)):
    dtr = (t1y2[i] - t1y1[i]) * 12 + t1m2[i] - t1m1[i]
    dt.append(dtr)

# V2 memory
t1v2 = list()
for i in range(len(variants)):
    t1v2r = t1v1[i] * 2**(dt[i]/t1Tmoor[i])
    t1v2.append(t1v2r)








b = list() #Name list ----------------------------------------------------------
for row in range(2, sheetname.max_row + 1):
    b.append(sheetname[row][1].value)
random.shuffle(b)
a = list(filter(None, b))


#answer1 = t1v2 # Answer to task 1, 2, 3 in order to search len variants --------------------------------------
b = t1v2


if len(b) < len(a): # Start logic expression ----------------------------------------
    Max = a
    Min = b
else: Max = b; Min = a
LenMax = len(Max)

for i in range(LenMax): # Start write to txt file
    if a == Max:
        bb = (i % len(Min))
        aa = i
    else:
        aa = (i % len(Min))
        bb = i
    name = "ЛР2_Закон Мура " + str(a[aa])
    my_file = '{}.txt'.format(name)
    with open(my_file, "w") as f:
        f.write("- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -   \n")
        f.write("ОМ ЛР 02. Закон Мура. Варіант" + "(" + str(i+1) + ")\n")
        f.write("Завдання 1. ------------------------- \n")
        f.write("Дано: \n")
        f.write(" Y1   M1  Y2   M2  V1  Tmoor \n")
        f.write(str(t1y1[bb]) +"  "+str(t1m1[bb]) +"  "+str(t1y2[bb]) +"  "+str(t1m2[bb]) +"  "+str(t1v1[bb]) +"  "+str(t1Tmoor[bb]) +"  "  + "\n \v")
        f.write(
                "В M1 місяці Y1 року флеш пристрій ємністю V1 коштував 70$.\n"
                "1. Знайти ємність V2, Мб, флеш пристрію такої же користувацької цінності\n"
                "в Y2 році M2 місяці.\n"
                "Вважаємо, що флеш-накопичувачі подвоюють свою ємність кожні Tmoor місяців.\n")



my_answer = '1Answer.txt'
for i in range(LenMax): # Write Answer -----------------------
    if a == Max:
        bb = (i % len(Min))
        aa = i
    else:
        aa = (i % len(Min))
        bb = i
    with open(my_answer, "a") as ma:
        ma.write(str(a[aa]) + " Answer Laba is: \n"  
        "Task 1 --- " + str(t1v2[bb]) + "\n"
        "Task 2 --- " + "\n"
        "Task 3 --- " + "\v")











