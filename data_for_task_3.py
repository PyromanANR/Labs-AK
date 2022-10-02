import openpyxl
book = openpyxl.load_workbook("table_task_3.xlsx", read_only=True)
sheet = book.worksheets[0]

variants = list()
for row in range(3, sheet.max_row):
    variants.append(sheet[row][0].value)

t3y1 = list()
for row in range(3, sheet.max_row):
    t3y1.append(sheet[row][1].value)

t3m1 = list()
for row in range(3, sheet.max_row):
    t3m1.append(sheet[row][2].value)

t3y2 = list()
for row in range(3, sheet.max_row):
    t3y2.append(sheet[row][3].value)

t3m2 = list()
for row in range(3, sheet.max_row):
    t3m2.append(sheet[row][4].value)

t3v1 = list()
for row in range(3, sheet.max_row):
    t3v1.append(sheet[row][5].value)

t3v2 = list()
for row in range(3, sheet.max_row):
    t3v2.append(sheet[row][6].value)

    
